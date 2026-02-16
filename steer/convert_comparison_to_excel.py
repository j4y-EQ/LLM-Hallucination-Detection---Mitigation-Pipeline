#!/usr/bin/env python
r"""
Convert Comparison Results CSV to Excel

Converts the comparison results CSV file from cross_experiment_analysis.py to a 
well-formatted Excel file with conditional formatting, colored headers, and 
improved readability.

Usage:
    python convert_comparison_to_excel.py <csv_file_path>
    python convert_comparison_to_excel.py --input "/home/jovyan/HaluEval/data/ITI/steering_experiment_qwen2.5_7b_3_percent/round3/cross_analysis/RUN_20260204_073337/comparison_results_20260204_073337.csv" --output results.xlsx

Features:
    - Automatically detects single or dual secondary mode
    - Color-coded valid/invalid configurations
    - Conditional formatting for pass/fail columns
    - Percentage formatting for rate columns
    - Frozen header row for easy scrolling
    - Auto-sized columns for optimal viewing
"""

import pandas as pd
import argparse
import os
import sys
from pathlib import Path


def convert_csv_to_excel(csv_path: str, output_path: str = None, use_formatting: bool = True) -> str:
    """
    Convert comparison results CSV to Excel format.
    
    Args:
        csv_path: Path to the input CSV file
        output_path: Optional path for output Excel file (default: same as CSV with .xlsx extension)
        use_formatting: Whether to apply conditional formatting (requires openpyxl)
    
    Returns:
        Path to the generated Excel file
    """
    
    # Validate input file
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    print(f"Reading CSV file: {csv_path}")
    
    # Read CSV
    df = pd.read_csv(csv_path)
    
    print(f"Loaded {len(df)} rows and {len(df.columns)} columns")
    
    # Determine output path
    if output_path is None:
        output_path = str(Path(csv_path).with_suffix('.xlsx'))
    
    # Detect mode (single vs dual secondary)
    has_dual_secondary = 'secondary_2_baseline_rate' in df.columns
    mode = "dual" if has_dual_secondary else "single"
    print(f"Detected mode: {mode} secondary")
    
    if use_formatting:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Write to Excel with basic pandas
            df.to_excel(output_path, index=False, engine='openpyxl')
            
            # Load workbook for formatting
            wb = load_workbook(output_path)
            ws = wb.active
            
            print("Applying formatting...")
            
            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            valid_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green
            invalid_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
            
            pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")  # Light green
            fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")  # Light red
            
            center_align = Alignment(horizontal="center", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Format header row
            for col_idx, col in enumerate(df.columns, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border
            
            # Format data rows
            for row_idx in range(2, len(df) + 2):
                # Get valid_config value for this row
                valid_config_col = df.columns.get_loc('valid_config') + 1
                valid_config_value = ws.cell(row=row_idx, column=valid_config_col).value
                
                for col_idx, col in enumerate(df.columns, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    
                    # Apply alignment
                    if col in ['k', 'alpha']:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                    
                    # Format rate columns as percentages
                    if 'rate' in col.lower() and col != 'max_allowed_rate':
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                    
                    # Format reduction columns with 2 decimal places
                    if 'reduction' in col.lower():
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00'
                    
                    # Format threshold columns as percentages
                    if 'max_allowed_rate' in col:
                        if cell.value is not None and isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00%'
                    
                    # Color-code valid_config column
                    if col == 'valid_config':
                        if valid_config_value == True:
                            cell.fill = valid_fill
                            cell.font = Font(bold=True, color="006100")
                        elif valid_config_value == False:
                            cell.fill = invalid_fill
                            cell.font = Font(bold=True, color="9C0006")
                        cell.alignment = center_align
                    
                    # Color-code pass columns
                    if col.endswith('_passes'):
                        if cell.value == True:
                            cell.fill = pass_fill
                            cell.font = Font(color="155724")
                        elif cell.value == False:
                            cell.fill = fail_fill
                            cell.font = Font(color="721C24")
                        cell.alignment = center_align
                    
                    # Highlight failure reasons
                    if col.endswith('_fail_reason') and cell.value:
                        cell.font = Font(italic=True, color="856404")
                        cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            
            # Auto-size columns
            for col_idx, col in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_idx)
                
                # Calculate max width
                max_length = len(str(col))
                for row_idx in range(2, min(len(df) + 2, 102)):  # Sample first 100 rows
                    try:
                        cell_value = ws.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    except:
                        pass
                
                # Set column width (add padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header row
            ws.freeze_panes = 'A2'
            
            # Save formatted workbook
            wb.save(output_path)
            
            print(f"✓ Formatted Excel file created: {output_path}")
            
        except ImportError:
            print("⚠ openpyxl not available - creating basic Excel file without formatting")
            df.to_excel(output_path, index=False)
            print(f"✓ Basic Excel file created: {output_path}")
    
    else:
        # Basic export without formatting
        df.to_excel(output_path, index=False)
        print(f"✓ Excel file created: {output_path}")
    
    # Print summary statistics
    print("\n" + "="*60)
    print("SUMMARY")
    print("="*60)
    print(f"Total configurations: {len(df)}")
    
    if 'valid_config' in df.columns:
        valid_count = df['valid_config'].sum()
        print(f"Valid configurations: {valid_count}")
        print(f"Invalid configurations: {len(df) - valid_count}")
    
    if 'primary_absolute_reduction' in df.columns:
        best_idx = df['primary_absolute_reduction'].idxmax()
        best_row = df.loc[best_idx]
        print(f"\nBest primary reduction:")
        print(f"  k={int(best_row['k'])}, alpha={best_row['alpha']:.1f}")
        print(f"  Absolute reduction: {best_row['primary_absolute_reduction']:.2f} pp")
        
        if 'valid_config' in df.columns:
            valid_df = df[df['valid_config'] == True]
            if not valid_df.empty:
                best_valid_idx = valid_df['primary_absolute_reduction'].idxmax()
                best_valid_row = valid_df.loc[best_valid_idx]
                print(f"\nBest valid configuration:")
                print(f"  k={int(best_valid_row['k'])}, alpha={best_valid_row['alpha']:.1f}")
                print(f"  Absolute reduction: {best_valid_row['primary_absolute_reduction']:.2f} pp")
    
    print("="*60 + "\n")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Convert comparison results CSV to formatted Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Convert with automatic output name
  python convert_comparison_to_excel.py comparison_results_20260131_130527.csv

  # Specify output file
  python convert_comparison_to_excel.py --input comparison.csv --output results.xlsx

  # Convert without formatting (basic Excel)
  python convert_comparison_to_excel.py --input comparison.csv --no-formatting
        """
    )
    
    parser.add_argument(
        'csv_file', nargs='?', default=None,
        help='Path to comparison results CSV file'
    )
    parser.add_argument(
        '--input', '-i', type=str, default=None,
        help='Path to comparison results CSV file (alternative to positional argument)'
    )
    parser.add_argument(
        '--output', '-o', type=str, default=None,
        help='Path to output Excel file (default: same as input with .xlsx extension)'
    )
    parser.add_argument(
        '--no-formatting', action='store_true',
        help='Skip conditional formatting (creates basic Excel file)'
    )
    
    args = parser.parse_args()
    
    # Determine input file
    csv_file = args.csv_file or args.input
    
    if not csv_file:
        parser.error("No input file specified. Use positional argument or --input flag.")
    
    try:
        output_file = convert_csv_to_excel(
            csv_file,
            output_path=args.output,
            use_formatting=not args.no_formatting
        )
        print(f"\n✓ Conversion complete!")
        print(f"Output: {output_file}")
        
    except Exception as e:
        print(f"\n✗ Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
